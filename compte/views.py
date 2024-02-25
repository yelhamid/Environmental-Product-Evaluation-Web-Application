from audioop import mul
from cmath import e
from operator import length_hint
from tkinter import CENTER, font
from turtle import color
from unicodedata import name
from django.utils.http import urlsafe_base64_decode ,urlsafe_base64_encode
from django.utils.encoding import force_bytes, force_text
from django.template.loader import render_to_string
from django.contrib.sites.shortcuts import get_current_site 
from projet import settings
from distutils.log import error
from django.contrib.auth.decorators import login_required
import email
from multiprocessing import context
from django.shortcuts import render,redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate , login , logout
from django.contrib import messages
from django.core.mail import send_mail, EmailMessage
from .token import generateToken
from .models import *
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import *
import openpyxl as xl
import xlwt
from django.http import HttpResponse
from django.contrib.auth.models import User


# Create your views here.








@login_required
def utiliser(request ):
    return render(request, 'compte/calcule.html',) 

def inscriptionPage(request):
    if request.method == 'POST':
        username = request.POST['username'] 
        firstname = request.POST['firstname'] 
        lastname = request.POST['lastname']  
        email = request.POST['email']
        password1 = request.POST['password1']
        password2 = request.POST['password2']
        if (len(username)==0 or len(firstname)==0 or len(lastname)==0 or len(email)==0 or len(password1)==0 or len(password2)==0):
            messages.error(request,'Il faut remplir tous les champs')
            return render(request,'compte/inscription.html',{'errorchamps':'is-invalid'})
        if User.objects.filter(username=username):
            messages.error(request, "nom d'utilisateur déjà existe")
            return render(request,'compte/inscription.html',{'erroruser':'is-invalid'}) 
        if User.objects.filter(email=email):
            messages.error(request, "Cet email a déjà un compte")
            return render(request,'compte/inscription.html',{'erroremail':'is-invalid'}) 
        if len(username)>10:
            messages.error(request, "Le nom d'utilisateur ne doit pas être plus de 10 caractères.")
            return render(request,'compte/inscription.html',{'erroruser':'is-invalid'}) 
        if len(username)<5:
            messages.error(request, "S'il vous plaît, le nom d'utilisateur doit comporter au moins 5 caractères.")
            return render(request,'compte/inscription.html',{'erroruser':'is-invalid'}) 
        if not username.isalnum():
            messages.error(request , "Le nom d'utilisateur doit être alphanumérique")
            return render(request,'compte/inscription.html',{'erroruser':'is-invalid'}) 
        if len(password1)<5:
            messages.error(request, "S'il vous plaît, le mot de passe doit comporter au moins 5 caractères.")
            return render(request,'compte/inscription.html',{'errorpass':'is-invalid'}) 
        if password1 !=password2:
            messages.error(request, 'Les deux mots de passe ne correspondent pas.')
            return render(request,'compte/inscription.html',{'errorpass':'is-invalid'}) 

        mon_utilisateur = User.objects.create_user(username , email , password1)
        mon_utilisateur.first_name = firstname
        mon_utilisateur.last_name = lastname
        mon_utilisateur.is_active = False
        mon_utilisateur.save()
        messages.success(request,'Votre compte a été créé avec succès, vérifiez votre boite mail et activez votre compte')

        #envoi d'email de bienvenu :
        subject = "Bienvenue sur notre application web"
        message = "Bienvenue "+ mon_utilisateur.first_name + " " + mon_utilisateur.last_name + "\nNous sommes heureux de vous compter parmi nous \n\n\n Merci. "
        from_email = settings.EMAIL_HOST_USER
        to_list = [mon_utilisateur.email]
        send_mail(subject , message ,from_email ,to_list , fail_silently=False)

        #Email de confirmation
        current_site = get_current_site(request)
        email_subject = "Confirmation de l'adresse email sur notre site web"
        messageConfirm = render_to_string("emailcomfirm.html", {
            "name" : mon_utilisateur.first_name,
            "lastname" :mon_utilisateur.last_name,
            'domain' : current_site.domain,
            'uid':urlsafe_base64_encode(force_bytes(mon_utilisateur.pk)),
            'token': generateToken.make_token(mon_utilisateur)
        })

        email = EmailMessage(
            email_subject,
            messageConfirm,
            settings.EMAIL_HOST_USER,
            [mon_utilisateur.email]
        )

        email.fail_silently = False
        email.send()
        

        return render(request,'compte/acces.html',{'susses':'alert-success'}) 
    
    return render(request,'compte/inscription.html')
 



def accesPage(request):
    if request.method == 'POST':
        username = request.POST['username'] 
        password = request.POST['password'] 
        user = authenticate(username=username,password=password)

        # my_user= User.objects.get(username=username)

        if user is not None:
            login(request, user)
            # user = User.objects.get(id=id_user)

            # prenom = user.first_name
            # nom = user.last_name
            # return render(request, 'compte/calcule.html', {'firstname':prenom , 'lastname':nom})
            return redirect('utiliser')
        
        # elif my_user.is_active == False:
        #     messages.error(request ,"Vous n'avez pas confirmez votre compte faite le avant de vous connecter merci!")
        #     return redirect('acces') 
        else:
            if (len(username)==0 or len(password)==0 ):
                messages.error(request,'il faut remplir tout les champs')
                return render(request,'compte/acces.html',{'classe':'is-invalid','danger':'alert-danger'})
            else:
                messages.error(request,"Nom d'utilisateur ou mot de passe est incorrect")
                return render(request,'compte/acces.html',{'classe':'is-invalid','danger':'alert-danger'})
    
    return render(request,'compte/acces.html')
    



 

@login_required
def logOut(request):
    logout(request)
    # messages.success(request,'vous avez été bien déconnecte')
    # return render(request,'compte/acces.html',{'susses':'alert-success'})
    return redirect('acces')


def activate(request,uidb64 ,token):
    try:
        uid = force_text(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except(TypeError,ValueError,OverflowError, User.DoesNotExist):
        user = None


    if user is not None and generateToken.check_token(user,token):
        user.is_active = True
        user.save()
        messages.success(request,"Votre compte a été activer, félicitations connectez-vous maintenant")
        return render(request,'compte/acces.html',{'susses':'alert-success'}) 
    else:
        messages.error(request, "Activation échoué !!! ")
        return render(request,'compte/acces.html',{'danger':'alert-danger'}) 


          
#  calcule sur la matiere :::::------------------------------------------------------------------------
@login_required
def matiere(request):
    if request.method == 'POST':
        
        test = request.POST.getlist('choix')
        choix= test[0]
        nom_fich = request.POST['nom_fich']
        multiplicator = request.POST['quantite']
        calc = multiplicator

        if choix=="Aa002":
            objs = Aa002.objects.all()
        elif choix=="Ab001":
            objs = Ab001.objects.all()
        elif choix=="Ac001":
            objs = Ac001.objects.all()
        elif choix=="Ac008":
            objs = Ac008.objects.all()
            

          
        
 
        if (len(nom_fich)==0):
            messages.error(request,'il faut remplir tout les champs')
            return render(request,'calcules/matiere.html',{'danger':'alert-danger'}) 

   

        mult =  multiplicator.replace(".","")
        
        
        ListS = list(mult)
        if  ListS[0]== "-":
            ListS.pop(0)
            mult = "".join(ListS)
        
        if mult.isnumeric():
            if float(multiplicator) < 0 :
                messages.error(request,'La quantité de matière doit être positif')
                return render(request,'calcules/matiere.html',{'danger':'alert-danger'}) 
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = choix
            

                                        
                #Remplir les cellule
                ws['A1'] = objs[0].a
                ws['A5'] = objs[2].a
                ws['A7'] = objs[3].a


                new = str(multiplicator)
                new_val = objs[0].e
                new_val = new_val.replace("1", new)
                ws['E1'] = new_val
                ws['E7'] = objs[3].e
        

                ws['H1'] = objs[0].h
                ws['H3'] = objs[1].h
                ws['H5'] = objs[2].h
                ws['H7'] = objs[3].h
                

                for i in range(9, 29):
                    x = i - 5           
                    ws['A'+str(i)]=objs[x].a
                    ws['H'+str(i)]=objs[x].h
                    
                    do =(objs[i-5].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(objs[i-5].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(calc)
                    ws['E'+str(i)]=resultat
                    
                    
                    

                # DECLARATIONS DE QUELQUES CELLULES FREQUEMMENT UTILISEES
                a1 = ws['A1']
                e1 = ws['E1']
                a5 = ws['A5']
                a7 = ws['A7']
                e7 = ws['E7']
                h7 = ws['H7']
                h1 = ws['H1']
                h3 = ws['H3']
                h5 = ws['H5']
                a28 = ws['A28']
                e28 = ws['E28']
                h28 = ws['H28']

                # BORDURES ET ALIGNEMENT DES CELLULES
                thin = Side(border_style="thin", color="000000")
                double = Side(border_style="double", color="000000")

                def centrage(cellule):
                    cellule.alignment = Alignment(horizontal="center", vertical="center")

                def vertivaly_centrage(cellule):
                    cellule.alignment = Alignment(vertical="center")

                def blue_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="228B22")

                def bleuet_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="90EE90")

                def grey_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="DDDDDD")

                for i in range(9, 29):
                    j = ws['E' + str(i)]
                    centrage(j)

                for i in range(9, 29):
                    j = ws['H' + str(i)]
                    centrage(j)

                a1.border = Border(top=double, left=double)
                blue_background(a1)
                vertivaly_centrage(a1)

                e1.border = Border(top=double)
                blue_background(e1)
                centrage(e1)

                h1.border = Border(top=double, left=thin, right=double, bottom=thin)
                blue_background(h1)
                vertivaly_centrage(h1)

                h3.border = Border(top=thin, left=thin, right=double, bottom=thin)
                blue_background(h3)
                vertivaly_centrage(h3)

                h5.border = Border(top=thin, left=thin, right=double, bottom=double)
                blue_background(h5)
                vertivaly_centrage(h5)

                a7.border = Border(top=double, left=double, bottom=double)
                blue_background(a7)
                centrage(a7)

                e7.border = Border(top=double, bottom=double)
                blue_background(e7)
                centrage(e7)

                h7.border = Border(top=double, right=double, bottom=double)
                blue_background(h7)
                centrage(h7)

                a5.border = Border(left=double, bottom=double)
                blue_background(a5)
                vertivaly_centrage(a5)

                for i in range(9, 28):
                    t = ws['A' + str(i)]
                    t.border = Border(left=double)
                    t = ws['H' + str(i)]
                    t.border = Border(right=double)

                k = 9
                while k < 29:
                    t = ws['A' + str(k)]
                    bleuet_background(t)
                    e = ws['E' + str(k)]
                    h = ws['H' + str(k)]
                    grey_background(e)
                    grey_background(h)
                    k += 2

                k = 10
                while k < 29:
                    t = ws['A' + str(k)]
                    blue_background(t)
                    k += 2

                a28.border = Border(left=double, bottom=double)
                e28.border = Border(bottom=double)
                h28.border = Border(right=double, bottom=double)

                # FUSION DES CELLULES

                ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=4)
                ws.merge_cells(start_row=1, start_column=5, end_row=4, end_column=7)
                ws.merge_cells(start_row=1, start_column=8, end_row=2, end_column=9)
                ws.merge_cells(start_row=3, start_column=8, end_row=4, end_column=9)
                ws.merge_cells(start_row=5, start_column=8, end_row=6, end_column=9)
                ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=7)
                ws.merge_cells(start_row=7, start_column=1, end_row=8, end_column=4)
                ws.merge_cells(start_row=7, start_column=5, end_row=8, end_column=7)
                ws.merge_cells(start_row=7, start_column=8, end_row=8, end_column=9)

                def fusion(n):
                    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=4)
                    ws.merge_cells(start_row=n, start_column=5, end_row=n, end_column=7)
                    ws.merge_cells(start_row=n, start_column=8, end_row=n, end_column=9)

                for i in range(9, 29):
                    fusion(i)

                # MISE EN PAGE DU TEXTE
                a1.font = Font(color="FFFFFF", bold=True, size=14)

                e1.font = Font(color="FFFFFF", bold=True, size=14)

                a7.font = Font(color="000000", bold=True, size=14)

                e7.font = Font(color="000000", bold=True, size=14)

                h7.font = Font(color="000000", bold=True, size=14)

                h1.font = Font(color="FFFFFF", bold=True, size=11)

                def titrer(cellule):
                    cellule.font = Font(color="FFFFFF", size=11)

                titrer(h3)
                titrer(h5)

                for i in range(9, 29):
                    t = ws['A' + str(i)]
                    titrer(t)
                    centrage(t)

                wb.save('/Users/lenovo/Desktop/'+nom_fich+'.xlsx')

                



                messages.success(request,"Votre document excel a été généré avec succes")
                return render(request,'calcules/matiere.html',{'susses':'alert-success'})  

        else:
            messages.error(request,"il faut saisir un nombre dans le champs du quantité de matière")
            return render(request,'calcules/matiere.html',{'danger':'alert-danger'}) 
    return render(request,'calcules/matiere.html') 

# Fin calcule sur la matiere ------------------------------------------------------------------------



# calcule sur les procedes :::------------------------------------------------------------------------
@login_required
def procedes(request):
    if request.method == 'POST':
        
        test = request.POST.getlist('choix')
        choix= test[0]
        nom_fich = request.POST['nom_fich']
        multiplicator = request.POST['quantite']
        calc = multiplicator

        if choix=="Ba043":
            objs = Ba043.objects.all()
        elif choix=="Ba044":
            objs = Ba044.objects.all()
        elif choix=="Ba045":
            objs = Ba045.objects.all()
        elif choix=="Bf004":
            objs = Bf004.objects.all()
            

          
        
 
        if (len(nom_fich)==0):
            messages.error(request,'il faut remplir tout les champs')
            return render(request,'calcules/procedes.html',{'danger':'alert-danger'}) 

   

        mult =  multiplicator.replace(".","")
        
        
        ListS = list(mult)
        if  ListS[0]== "-":
            ListS.pop(0)
            mult = "".join(ListS)
        
        if mult.isnumeric():
            if float(multiplicator) < 0 :
                messages.error(request,'La quantité de matière doit être positif')
                return render(request,'calcules/procedes.html',{'danger':'alert-danger'}) 
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = choix
            

                                        
                #Remplir les cellule
                ws['A1'] = objs[0].a
                ws['A5'] = objs[2].a
                ws['A7'] = objs[3].a


                new = str(multiplicator)
                new_val = objs[0].e
                new_val = new_val.replace("1", new)
                ws['E1'] = new_val
                ws['E7'] = objs[3].e
        

                ws['H1'] = objs[0].h
                ws['H3'] = objs[1].h
                ws['H5'] = objs[2].h
                ws['H7'] = objs[3].h
                

                for i in range(9, 29):
                    x = i - 5           
                    ws['A'+str(i)]=objs[x].a
                    ws['H'+str(i)]=objs[x].h
                    do =(objs[i-5].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(objs[i-5].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(calc)
                    ws['E'+str(i)]=resultat
                    
                    
                    

                # DECLARATIONS DE QUELQUES CELLULES FREQUEMMENT UTILISEES
                a1 = ws['A1']
                e1 = ws['E1']
                a5 = ws['A5']
                a7 = ws['A7']
                e7 = ws['E7']
                h7 = ws['H7']
                h1 = ws['H1']
                h3 = ws['H3']
                h5 = ws['H5']
                a28 = ws['A28']
                e28 = ws['E28']
                h28 = ws['H28']

                # BORDURES ET ALIGNEMENT DES CELLULES
                thin = Side(border_style="thin", color="000000")
                double = Side(border_style="double", color="000000")

                def centrage(cellule):
                    cellule.alignment = Alignment(horizontal="center", vertical="center")

                def vertivaly_centrage(cellule):
                    cellule.alignment = Alignment(vertical="center")

                def blue_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="228B22")

                def bleuet_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="90EE90")

                def grey_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="DDDDDD")

                for i in range(9, 29):
                    j = ws['E' + str(i)]
                    centrage(j)

                for i in range(9, 29):
                    j = ws['H' + str(i)]
                    centrage(j)

                a1.border = Border(top=double, left=double)
                blue_background(a1)
                vertivaly_centrage(a1)

                e1.border = Border(top=double)
                blue_background(e1)
                centrage(e1)

                h1.border = Border(top=double, left=thin, right=double, bottom=thin)
                blue_background(h1)
                vertivaly_centrage(h1)

                h3.border = Border(top=thin, left=thin, right=double, bottom=thin)
                blue_background(h3)
                vertivaly_centrage(h3)

                h5.border = Border(top=thin, left=thin, right=double, bottom=double)
                blue_background(h5)
                vertivaly_centrage(h5)

                a7.border = Border(top=double, left=double, bottom=double)
                blue_background(a7)
                centrage(a7)

                e7.border = Border(top=double, bottom=double)
                blue_background(e7)
                centrage(e7)

                h7.border = Border(top=double, right=double, bottom=double)
                blue_background(h7)
                centrage(h7)

                a5.border = Border(left=double, bottom=double)
                blue_background(a5)
                vertivaly_centrage(a5)

                for i in range(9, 28):
                    t = ws['A' + str(i)]
                    t.border = Border(left=double)
                    t = ws['H' + str(i)]
                    t.border = Border(right=double)

                k = 9
                while k < 29:
                    t = ws['A' + str(k)]
                    bleuet_background(t)
                    e = ws['E' + str(k)]
                    h = ws['H' + str(k)]
                    grey_background(e)
                    grey_background(h)
                    k += 2

                k = 10
                while k < 29:
                    t = ws['A' + str(k)]
                    blue_background(t)
                    k += 2

                a28.border = Border(left=double, bottom=double)
                e28.border = Border(bottom=double)
                h28.border = Border(right=double, bottom=double)

                # FUSION DES CELLULES

                ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=4)
                ws.merge_cells(start_row=1, start_column=5, end_row=4, end_column=7)
                ws.merge_cells(start_row=1, start_column=8, end_row=2, end_column=9)
                ws.merge_cells(start_row=3, start_column=8, end_row=4, end_column=9)
                ws.merge_cells(start_row=5, start_column=8, end_row=6, end_column=9)
                ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=7)
                ws.merge_cells(start_row=7, start_column=1, end_row=8, end_column=4)
                ws.merge_cells(start_row=7, start_column=5, end_row=8, end_column=7)
                ws.merge_cells(start_row=7, start_column=8, end_row=8, end_column=9)

                def fusion(n):
                    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=4)
                    ws.merge_cells(start_row=n, start_column=5, end_row=n, end_column=7)
                    ws.merge_cells(start_row=n, start_column=8, end_row=n, end_column=9)

                for i in range(9, 29):
                    fusion(i)

                # MISE EN PAGE DU TEXTE
                a1.font = Font(color="FFFFFF", bold=True, size=14)

                e1.font = Font(color="FFFFFF", bold=True, size=14)

                a7.font = Font(color="000000", bold=True, size=14)

                e7.font = Font(color="000000", bold=True, size=14)

                h7.font = Font(color="000000", bold=True, size=14)

                h1.font = Font(color="FFFFFF", bold=True, size=11)

                def titrer(cellule):
                    cellule.font = Font(color="FFFFFF", size=11)

                titrer(h3)
                titrer(h5)

                for i in range(9, 29):
                    t = ws['A' + str(i)]
                    titrer(t)
                    centrage(t)

                wb.save('/Users/lenovo/Desktop/'+nom_fich+'.xlsx')

                



                messages.success(request,"Votre document excel a été généré avec succes")
                return render(request,'calcules/procedes.html',{'susses':'alert-success'})  

        else:
            messages.error(request,"il faut saisir un nombre dans le champs du quantité de matière")
            return render(request,'calcules/procedes.html',{'danger':'alert-danger'}) 
    return render(request,'calcules/procedes.html') 

# Fin calcule sur procedes :__________________________________________________________________________



# calcule sur energie :: ____________________________________________________________________________
@login_required
def energie(request):
    if request.method == 'POST':
        
        test = request.POST.getlist('choix')
        choix= test[0]
        nom_fich = request.POST['nom_fich']
        multiplicator = request.POST['quantite']
        calc = multiplicator

        if choix=="Ca001":
            objs = Ca001.objects.all()
        elif choix=="Cb001":
            objs = Cb001.objects.all()
      
            

          
        
 
        if (len(nom_fich)==0):
            messages.error(request,'il faut remplir tout les champs')
            return render(request,'calcules/energie.html',{'danger':'alert-danger'}) 

   

        mult =  multiplicator.replace(".","")
        
        
        ListS = list(mult)
        if  ListS[0]== "-":
            ListS.pop(0)
            mult = "".join(ListS)
        
        if mult.isnumeric():
            if float(multiplicator) < 0 :
                messages.error(request,'La quantité de matière doit être positif')
                return render(request,'calcules/energie.html',{'danger':'alert-danger'}) 
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = choix
            

                                        
                #Remplir les cellule
                ws['A1'] = objs[0].a
                ws['A5'] = objs[2].a
                ws['A7'] = objs[3].a


                new = str(multiplicator)
                new_val = objs[0].e
                new_val = new_val.replace("1", new)
                ws['E1'] = new_val
                ws['E7'] = objs[3].e
        

                ws['H1'] = objs[0].h
                ws['H3'] = objs[1].h
                ws['H5'] = objs[2].h
                ws['H7'] = objs[3].h
                

                for i in range(9, 29):
                    x = i - 5           
                    ws['A'+str(i)]=objs[x].a
                    ws['H'+str(i)]=objs[x].h
                    do =(objs[i-5].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(objs[i-5].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(calc)
                    ws['E'+str(i)]=resultat
                    
                    
                    

                # DECLARATIONS DE QUELQUES CELLULES FREQUEMMENT UTILISEES
                a1 = ws['A1']
                e1 = ws['E1']
                a5 = ws['A5']
                a7 = ws['A7']
                e7 = ws['E7']
                h7 = ws['H7']
                h1 = ws['H1']
                h3 = ws['H3']
                h5 = ws['H5']
                a28 = ws['A28']
                e28 = ws['E28']
                h28 = ws['H28']

                # BORDURES ET ALIGNEMENT DES CELLULES
                thin = Side(border_style="thin", color="000000")
                double = Side(border_style="double", color="000000")

                def centrage(cellule):
                    cellule.alignment = Alignment(horizontal="center", vertical="center")

                def vertivaly_centrage(cellule):
                    cellule.alignment = Alignment(vertical="center")

                def blue_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="228B22")

                def bleuet_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="90EE90")

                def grey_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="DDDDDD")

                for i in range(9, 29):
                    j = ws['E' + str(i)]
                    centrage(j)

                for i in range(9, 29):
                    j = ws['H' + str(i)]
                    centrage(j)

                a1.border = Border(top=double, left=double)
                blue_background(a1)
                vertivaly_centrage(a1)

                e1.border = Border(top=double)
                blue_background(e1)
                centrage(e1)

                h1.border = Border(top=double, left=thin, right=double, bottom=thin)
                blue_background(h1)
                vertivaly_centrage(h1)

                h3.border = Border(top=thin, left=thin, right=double, bottom=thin)
                blue_background(h3)
                vertivaly_centrage(h3)

                h5.border = Border(top=thin, left=thin, right=double, bottom=double)
                blue_background(h5)
                vertivaly_centrage(h5)

                a7.border = Border(top=double, left=double, bottom=double)
                blue_background(a7)
                centrage(a7)

                e7.border = Border(top=double, bottom=double)
                blue_background(e7)
                centrage(e7)

                h7.border = Border(top=double, right=double, bottom=double)
                blue_background(h7)
                centrage(h7)

                a5.border = Border(left=double, bottom=double)
                blue_background(a5)
                vertivaly_centrage(a5)

                for i in range(9, 28):
                    t = ws['A' + str(i)]
                    t.border = Border(left=double)
                    t = ws['H' + str(i)]
                    t.border = Border(right=double)

                k = 9
                while k < 29:
                    t = ws['A' + str(k)]
                    bleuet_background(t)
                    e = ws['E' + str(k)]
                    h = ws['H' + str(k)]
                    grey_background(e)
                    grey_background(h)
                    k += 2

                k = 10
                while k < 29:
                    t = ws['A' + str(k)]
                    blue_background(t)
                    k += 2

                a28.border = Border(left=double, bottom=double)
                e28.border = Border(bottom=double)
                h28.border = Border(right=double, bottom=double)

                # FUSION DES CELLULES

                ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=4)
                ws.merge_cells(start_row=1, start_column=5, end_row=4, end_column=7)
                ws.merge_cells(start_row=1, start_column=8, end_row=2, end_column=9)
                ws.merge_cells(start_row=3, start_column=8, end_row=4, end_column=9)
                ws.merge_cells(start_row=5, start_column=8, end_row=6, end_column=9)
                ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=7)
                ws.merge_cells(start_row=7, start_column=1, end_row=8, end_column=4)
                ws.merge_cells(start_row=7, start_column=5, end_row=8, end_column=7)
                ws.merge_cells(start_row=7, start_column=8, end_row=8, end_column=9)

                def fusion(n):
                    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=4)
                    ws.merge_cells(start_row=n, start_column=5, end_row=n, end_column=7)
                    ws.merge_cells(start_row=n, start_column=8, end_row=n, end_column=9)

                for i in range(9, 29):
                    fusion(i)

                # MISE EN PAGE DU TEXTE
                a1.font = Font(color="FFFFFF", bold=True, size=14)

                e1.font = Font(color="FFFFFF", bold=True, size=14)

                a7.font = Font(color="000000", bold=True, size=14)

                e7.font = Font(color="000000", bold=True, size=14)

                h7.font = Font(color="000000", bold=True, size=14)

                h1.font = Font(color="FFFFFF", bold=True, size=11)

                def titrer(cellule):
                    cellule.font = Font(color="FFFFFF", size=11)

                titrer(h3)
                titrer(h5)

                for i in range(9, 29):
                    t = ws['A' + str(i)]
                    titrer(t)
                    centrage(t)

                wb.save('/Users/lenovo/Desktop/'+nom_fich+'.xlsx')

                



                messages.success(request,"Votre document excel a été généré avec succes")
                return render(request,'calcules/energie.html',{'susses':'alert-success'})  

        else:
            messages.error(request,"il faut saisir un nombre dans le champs du quantité de matière")
            return render(request,'calcules/energie.html',{'danger':'alert-danger'}) 
    return render(request,'calcules/energie.html') 

# FIn calcule sur energie ::_____________________________________________________________________________



#  calcule sur Fin de vie :: ______________________________________________________________________
@login_required
def fin_vie(request):
    if request.method == 'POST':
        
        test = request.POST.getlist('choix')
        choix= test[0]
        nom_fich = request.POST['nom_fich']
        multiplicator = request.POST['quantite']
        calc = multiplicator

        if choix=="Ez004":
            objs = Ez004.objects.all()
        elif choix=="Ez005":
            objs = Ez005.objects.all()
        elif choix=="Ez007":
            objs = Ez007.objects.all()
      
 
        if (len(nom_fich)==0):
            messages.error(request,'il faut remplir tout les champs')
            return render(request,'calcules/fin_vie.html',{'danger':'alert-danger'}) 

   

        mult =  multiplicator.replace(".","")
        
        
        ListS = list(mult)
        if  ListS[0]== "-":
            ListS.pop(0)
            mult = "".join(ListS)
        
        if mult.isnumeric():
            if float(multiplicator) < 0 :
                messages.error(request,'La quantité de matière doit être positif')
                return render(request,'calcules/fin_vie.html',{'danger':'alert-danger'}) 
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = choix
            

                                        
                #Remplir les cellule
                ws['A1'] = objs[0].a
                ws['A5'] = objs[2].a
                ws['A7'] = objs[3].a


                new = str(multiplicator)
                new_val = objs[0].e
                new_val = new_val.replace("1", new)
                ws['E1'] = new_val
                ws['E7'] = objs[3].e
        

                ws['H1'] = objs[0].h
                ws['H3'] = objs[1].h
                ws['H5'] = objs[2].h
                ws['H7'] = objs[3].h
                

                for i in range(9, 29):
                    x = i - 5           
                    ws['A'+str(i)]=objs[x].a
                    ws['H'+str(i)]=objs[x].h
                    do =(objs[i-5].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(objs[i-5].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(calc)
                    ws['E'+str(i)]=resultat
                    
                    
                    

                # DECLARATIONS DE QUELQUES CELLULES FREQUEMMENT UTILISEES
                a1 = ws['A1']
                e1 = ws['E1']
                a5 = ws['A5']
                a7 = ws['A7']
                e7 = ws['E7']
                h7 = ws['H7']
                h1 = ws['H1']
                h3 = ws['H3']
                h5 = ws['H5']
                a28 = ws['A28']
                e28 = ws['E28']
                h28 = ws['H28']

                # BORDURES ET ALIGNEMENT DES CELLULES
                thin = Side(border_style="thin", color="000000")
                double = Side(border_style="double", color="000000")

                def centrage(cellule):
                    cellule.alignment = Alignment(horizontal="center", vertical="center")

                def vertivaly_centrage(cellule):
                    cellule.alignment = Alignment(vertical="center")

                def blue_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="228B22")

                def bleuet_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="90EE90")

                def grey_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="DDDDDD")

                for i in range(9, 29):
                    j = ws['E' + str(i)]
                    centrage(j)

                for i in range(9, 29):
                    j = ws['H' + str(i)]
                    centrage(j)

                a1.border = Border(top=double, left=double)
                blue_background(a1)
                vertivaly_centrage(a1)

                e1.border = Border(top=double)
                blue_background(e1)
                centrage(e1)

                h1.border = Border(top=double, left=thin, right=double, bottom=thin)
                blue_background(h1)
                vertivaly_centrage(h1)

                h3.border = Border(top=thin, left=thin, right=double, bottom=thin)
                blue_background(h3)
                vertivaly_centrage(h3)

                h5.border = Border(top=thin, left=thin, right=double, bottom=double)
                blue_background(h5)
                vertivaly_centrage(h5)

                a7.border = Border(top=double, left=double, bottom=double)
                blue_background(a7)
                centrage(a7)

                e7.border = Border(top=double, bottom=double)
                blue_background(e7)
                centrage(e7)

                h7.border = Border(top=double, right=double, bottom=double)
                blue_background(h7)
                centrage(h7)

                a5.border = Border(left=double, bottom=double)
                blue_background(a5)
                vertivaly_centrage(a5)

                for i in range(9, 28):
                    t = ws['A' + str(i)]
                    t.border = Border(left=double)
                    t = ws['H' + str(i)]
                    t.border = Border(right=double)

                k = 9
                while k < 29:
                    t = ws['A' + str(k)]
                    bleuet_background(t)
                    e = ws['E' + str(k)]
                    h = ws['H' + str(k)]
                    grey_background(e)
                    grey_background(h)
                    k += 2

                k = 10
                while k < 29:
                    t = ws['A' + str(k)]
                    blue_background(t)
                    k += 2

                a28.border = Border(left=double, bottom=double)
                e28.border = Border(bottom=double)
                h28.border = Border(right=double, bottom=double)

                # FUSION DES CELLULES

                ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=4)
                ws.merge_cells(start_row=1, start_column=5, end_row=4, end_column=7)
                ws.merge_cells(start_row=1, start_column=8, end_row=2, end_column=9)
                ws.merge_cells(start_row=3, start_column=8, end_row=4, end_column=9)
                ws.merge_cells(start_row=5, start_column=8, end_row=6, end_column=9)
                ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=7)
                ws.merge_cells(start_row=7, start_column=1, end_row=8, end_column=4)
                ws.merge_cells(start_row=7, start_column=5, end_row=8, end_column=7)
                ws.merge_cells(start_row=7, start_column=8, end_row=8, end_column=9)

                def fusion(n):
                    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=4)
                    ws.merge_cells(start_row=n, start_column=5, end_row=n, end_column=7)
                    ws.merge_cells(start_row=n, start_column=8, end_row=n, end_column=9)

                for i in range(9, 29):
                    fusion(i)

                # MISE EN PAGE DU TEXTE
                a1.font = Font(color="FFFFFF", bold=True, size=14)

                e1.font = Font(color="FFFFFF", bold=True, size=14)

                a7.font = Font(color="000000", bold=True, size=14)

                e7.font = Font(color="000000", bold=True, size=14)

                h7.font = Font(color="000000", bold=True, size=14)

                h1.font = Font(color="FFFFFF", bold=True, size=11)

                def titrer(cellule):
                    cellule.font = Font(color="FFFFFF", size=11)

                titrer(h3)
                titrer(h5)

                for i in range(9, 29):
                    t = ws['A' + str(i)]
                    titrer(t)
                    centrage(t)

                wb.save('/Users/lenovo/Desktop/'+nom_fich+'.xlsx')

                



                messages.success(request,"Votre document excel a été généré avec succes")
                return render(request,'calcules/fin_vie.html',{'susses':'alert-success'})  

        else:
            messages.error(request,"il faut saisir un nombre dans le champs du quantité de matière")
            return render(request,'calcules/fin_vie.html',{'danger':'alert-danger'}) 
    return render(request,'calcules/fin_vie.html') 

# Fin calcule sur fin de vie __________________________________________________________________________



#  calcule sur le transport :: ________________________________________________________________________
@login_required
def transport(request):
    if request.method == 'POST':
        
        test = request.POST.getlist('choix')
        choix= test[0]
        nom_fich = request.POST['nom_fich']
        multiplicator = request.POST['quantite']
        calc = multiplicator

        nom_fich = request.POST['nom_fich']
        multiplicator = request.POST['quantite']

        if choix=="Da002":
            objs = Da002.objects.all()
        elif choix=="Db001":
            objs = Db001.objects.all()
        elif choix=="Db002":
            objs = Db002.objects.all()
        elif choix=="Dc001":
            objs = Dc001.objects.all()
        elif choix=="Dd001":
            objs = Dd001.objects.all()
        elif choix=="Dz001":
            objs = Dz001.objects.all()
        elif choix=="Dz002":
            objs = Dz002.objects.all()
        elif choix=="Dz003":
            objs = Dz003.objects.all()
        elif choix=="Dz004":
            objs = Dz004.objects.all()
      
 
        if (len(nom_fich)==0):
            messages.error(request,'il faut remplir tout les champs')
            return render(request,'calcules/transport.html',{'danger':'alert-danger'}) 

   

        mult =  multiplicator.replace(".","")
        
        
        ListS = list(mult)
        if  ListS[0]== "-":
            ListS.pop(0)
            mult = "".join(ListS)
        
        if mult.isnumeric():
            if float(multiplicator) < 0 :
                messages.error(request,'La quantité de matière doit être positif')
                return render(request,'calcules/transport.html',{'danger':'alert-danger'}) 
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = choix
            

                                        
                #Remplir les cellule
                ws['A1'] = objs[0].a
                ws['A5'] = objs[2].a
                ws['A7'] = objs[3].a


                new = str(multiplicator)
                new_val = objs[0].e
                new_val = new_val.replace("1", new)
                ws['E1'] = new_val
                ws['E7'] = objs[3].e
        

                ws['H1'] = objs[0].h
                ws['H3'] = objs[1].h
                ws['H5'] = objs[2].h
                ws['H7'] = objs[3].h
                

                for i in range(9, 29):
                    x = i - 5           
                    ws['A'+str(i)]=objs[x].a
                    ws['H'+str(i)]=objs[x].h
                    do =(objs[i-5].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(objs[i-5].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(calc)
                    ws['E'+str(i)]=resultat
                    
                    
                    

                # DECLARATIONS DE QUELQUES CELLULES FREQUEMMENT UTILISEES
                a1 = ws['A1']
                e1 = ws['E1']
                a5 = ws['A5']
                a7 = ws['A7']
                e7 = ws['E7']
                h7 = ws['H7']
                h1 = ws['H1']
                h3 = ws['H3']
                h5 = ws['H5']
                a28 = ws['A28']
                e28 = ws['E28']
                h28 = ws['H28']

                # BORDURES ET ALIGNEMENT DES CELLULES
                thin = Side(border_style="thin", color="000000")
                double = Side(border_style="double", color="000000")

                def centrage(cellule):
                    cellule.alignment = Alignment(horizontal="center", vertical="center")

                def vertivaly_centrage(cellule):
                    cellule.alignment = Alignment(vertical="center")

                def blue_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="228B22")

                def bleuet_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="90EE90")

                def grey_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="DDDDDD")

                for i in range(9, 29):
                    j = ws['E' + str(i)]
                    centrage(j)

                for i in range(9, 29):
                    j = ws['H' + str(i)]
                    centrage(j)

                a1.border = Border(top=double, left=double)
                blue_background(a1)
                vertivaly_centrage(a1)

                e1.border = Border(top=double)
                blue_background(e1)
                centrage(e1)

                h1.border = Border(top=double, left=thin, right=double, bottom=thin)
                blue_background(h1)
                vertivaly_centrage(h1)

                h3.border = Border(top=thin, left=thin, right=double, bottom=thin)
                blue_background(h3)
                vertivaly_centrage(h3)

                h5.border = Border(top=thin, left=thin, right=double, bottom=double)
                blue_background(h5)
                vertivaly_centrage(h5)

                a7.border = Border(top=double, left=double, bottom=double)
                blue_background(a7)
                centrage(a7)

                e7.border = Border(top=double, bottom=double)
                blue_background(e7)
                centrage(e7)

                h7.border = Border(top=double, right=double, bottom=double)
                blue_background(h7)
                centrage(h7)

                a5.border = Border(left=double, bottom=double)
                blue_background(a5)
                vertivaly_centrage(a5)

                for i in range(9, 28):
                    t = ws['A' + str(i)]
                    t.border = Border(left=double)
                    t = ws['H' + str(i)]
                    t.border = Border(right=double)

                k = 9
                while k < 29:
                    t = ws['A' + str(k)]
                    bleuet_background(t)
                    e = ws['E' + str(k)]
                    h = ws['H' + str(k)]
                    grey_background(e)
                    grey_background(h)
                    k += 2

                k = 10
                while k < 29:
                    t = ws['A' + str(k)]
                    blue_background(t)
                    k += 2

                a28.border = Border(left=double, bottom=double)
                e28.border = Border(bottom=double)
                h28.border = Border(right=double, bottom=double)

                # FUSION DES CELLULES

                ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=4)
                ws.merge_cells(start_row=1, start_column=5, end_row=4, end_column=7)
                ws.merge_cells(start_row=1, start_column=8, end_row=2, end_column=9)
                ws.merge_cells(start_row=3, start_column=8, end_row=4, end_column=9)
                ws.merge_cells(start_row=5, start_column=8, end_row=6, end_column=9)
                ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=7)
                ws.merge_cells(start_row=7, start_column=1, end_row=8, end_column=4)
                ws.merge_cells(start_row=7, start_column=5, end_row=8, end_column=7)
                ws.merge_cells(start_row=7, start_column=8, end_row=8, end_column=9)

                def fusion(n):
                    ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=4)
                    ws.merge_cells(start_row=n, start_column=5, end_row=n, end_column=7)
                    ws.merge_cells(start_row=n, start_column=8, end_row=n, end_column=9)

                for i in range(9, 29):
                    fusion(i)

                # MISE EN PAGE DU TEXTE
                a1.font = Font(color="FFFFFF", bold=True, size=14)

                e1.font = Font(color="FFFFFF", bold=True, size=14)

                a7.font = Font(color="000000", bold=True, size=14)

                e7.font = Font(color="000000", bold=True, size=14)

                h7.font = Font(color="000000", bold=True, size=14)

                h1.font = Font(color="FFFFFF", bold=True, size=11)

                def titrer(cellule):
                    cellule.font = Font(color="FFFFFF", size=11)

                titrer(h3)
                titrer(h5)

                for i in range(9, 29):
                    t = ws['A' + str(i)]
                    titrer(t)
                    centrage(t)

                wb.save('/Users/lenovo/Desktop/'+nom_fich+'.xlsx')

                



                messages.success(request,"Votre document excel a été généré avec succes")
                return render(request,'calcules/transport.html',{'susses':'alert-success'})  

        else:
            messages.error(request,"il faut saisir un nombre dans le champs du quantité de matière")
            return render(request,'calcules/transport.html',{'danger':'alert-danger'}) 
    return render(request,'calcules/transport.html') 

# fin calcule sur transport::____________________________________________________________________________________


# Generer un recap:___________________________________________________________________________________________
@login_required
def recap(request):
    if request.method == 'POST':
        
        nom_fich = request.POST['nom_fich']
        multiplicator = request.POST['quantite']
        calc = multiplicator
        multip = request.POST['distance']
        ca = multip

        choix_matiere = request.POST.getlist('choix_matiere')
        matiere= choix_matiere[0]
        choix_transport = request.POST.getlist('choix_transport')
        transport= choix_transport[0]
        choix_procedes = request.POST.getlist('choix_procedes')
        procedes= choix_procedes[0]
        choix_energie = request.POST.getlist('choix_energie')
        energie= choix_energie[0]
        choix_fin_vie = request.POST.getlist('choix_fin_vie')
        fin_vie= choix_fin_vie[0]

        
        if matiere=="Aa002":
            tab_matiere = Aa002.objects.all()
        elif matiere=="Ab001":
            tab_matiere = Ab001.objects.all()
        elif matiere=="Ac001":
            tab_matiere = Ac001.objects.all()
        elif matiere=="Ac008":
            tab_matiere = Ac008.objects.all()

        

        if transport=="Da002":
            tab_transport = Da002.objects.all()
        elif transport=="Db001":
            tab_transport = Db001.objects.all()
        elif transport=="Db002":
            tab_transport = Db002.objects.all()
        elif transport=="Dc001":
            tab_transport = Dc001.objects.all()
        elif transport=="Dd001":
            tab_transport = Dd001.objects.all()
        elif transport=="Dz001":
            tab_transport = Dz001.objects.all()
        elif transport=="Dz002":
            tab_transport = Dz002.objects.all()
        elif transport=="Dz003":
            tab_transport = Dz003.objects.all()
        elif transport=="Dz004":
            tab_transport = Dz004.objects.all()
      
        
        if procedes=="Ba043":
            tab_procedes = Ba043.objects.all()
        elif procedes=="Ba044":
            tab_procedes = Ba044.objects.all()
        elif procedes=="Ba045":
            tab_procedes = Ba045.objects.all()
        elif procedes=="Bf004":
            tab_procedes = Bf004.objects.all()

         
        if energie=="Ca001":
            tab_energie = Ca001.objects.all()
        elif energie=="Cb001":
            tab_energie = Cb001.objects.all()

        
        if fin_vie=="Ez004":
            tab_fin_vie = Ez004.objects.all()
        elif fin_vie=="Ez005":
            tab_fin_vie = Ez005.objects.all()
        elif fin_vie=="Ez007":
            tab_fin_vie = Ez007.objects.all()


        
 
        if (len(nom_fich)==0):
            messages.error(request,'il faut remplir tout les champs')
            return render(request,'calcules/recap.html',{'danger':'alert-danger'}) 

   

        mult =  multiplicator.replace(".","")
        m = multip.replace(".","")
        
        ListS = list(mult)
        if  ListS[0]== "-":
            ListS.pop(0)
            mult = "".join(ListS)
        
        Liste = list(m)
        if  Liste[0]== "-":
            Liste.pop(0)
            m = "".join(Liste)

        
        
        if mult.isnumeric() and m.isnumeric():
            if float(multiplicator) < 0 :
                messages.error(request,'La quantité de matière et la distance  doit être positifs')
                return render(request,'calcules/recap.html',{'danger':'alert-danger'}) 
            
            elif float(multip) < 0 :
                messages.error(request,'La distance  doit être positifs')
                return render(request,'calcules/recap.html',{'danger':'alert-danger'}) 
            else:

                # calcule sur matiere
                wb = Workbook()
                ws_matiere = wb.active
                ws_matiere.title = "Matière"
                                    
                #Remplir les cellule
                ws_matiere['A1'] = tab_matiere[0].a
                ws_matiere['A5'] = tab_matiere[2].a
                ws_matiere['A7'] = tab_matiere[3].a


                new = str(multiplicator)
                new_val = tab_matiere[0].e
                new_val = new_val.replace("1", new)
                ws_matiere['E1'] = new_val
                ws_matiere['E7'] = tab_matiere[3].e
        

                ws_matiere['H1'] = tab_matiere[0].h
                ws_matiere['H3'] = tab_matiere[1].h
                ws_matiere['H5'] = tab_matiere[2].h
                ws_matiere['H7'] = tab_matiere[3].h
                

                for i in range(9, 29):
                    x = i - 5           
                    ws_matiere['A'+str(i)]=tab_matiere[x].a
                    ws_matiere['H'+str(i)]=tab_matiere[x].h
                    do =(tab_matiere[i-5].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(tab_matiere[i-5].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(calc)
                    ws_matiere['E'+str(i)]=resultat
                    
                    
                    

                # DECLARATIONS DE QUELQUES CELLULES FREQUEMMENT UTILISEES
                a1 = ws_matiere['A1']
                e1 = ws_matiere['E1']
                a5 = ws_matiere['A5']
                a7 = ws_matiere['A7']
                e7 = ws_matiere['E7']
                h7 = ws_matiere['H7']
                h1 = ws_matiere['H1']
                h3 = ws_matiere['H3']
                h5 = ws_matiere['H5']
                a28 = ws_matiere['A28']
                e28 = ws_matiere['E28']
                h28 = ws_matiere['H28']

                # BORDURES ET ALIGNEMENT DES CELLULES
                thin = Side(border_style="thin", color="000000")
                double = Side(border_style="double", color="000000")

                def centrage(cellule):
                    cellule.alignment = Alignment(horizontal="center", vertical="center")

                def vertivaly_centrage(cellule):
                    cellule.alignment = Alignment(vertical="center")

                def blue_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="228B22")

                def bleuet_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="90EE90")

                def grey_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="DDDDDD")

                for i in range(9, 29):
                    j = ws_matiere['E' + str(i)]
                    centrage(j)

                for i in range(9, 29):
                    j = ws_matiere['H' + str(i)]
                    centrage(j)

                a1.border = Border(top=double, left=double)
                blue_background(a1)
                vertivaly_centrage(a1)

                e1.border = Border(top=double)
                blue_background(e1)
                centrage(e1)

                h1.border = Border(top=double, left=thin, right=double, bottom=thin)
                blue_background(h1)
                vertivaly_centrage(h1)

                h3.border = Border(top=thin, left=thin, right=double, bottom=thin)
                blue_background(h3)
                vertivaly_centrage(h3)

                h5.border = Border(top=thin, left=thin, right=double, bottom=double)
                blue_background(h5)
                vertivaly_centrage(h5)

                a7.border = Border(top=double, left=double, bottom=double)
                blue_background(a7)
                centrage(a7)

                e7.border = Border(top=double, bottom=double)
                blue_background(e7)
                centrage(e7)

                h7.border = Border(top=double, right=double, bottom=double)
                blue_background(h7)
                centrage(h7)

                a5.border = Border(left=double, bottom=double)
                blue_background(a5)
                vertivaly_centrage(a5)

                for i in range(9, 28):
                    t = ws_matiere['A' + str(i)]
                    t.border = Border(left=double)
                    t = ws_matiere['H' + str(i)]
                    t.border = Border(right=double)

                k = 9
                while k < 29:
                    t = ws_matiere['A' + str(k)]
                    bleuet_background(t)
                    e = ws_matiere['E' + str(k)]
                    h = ws_matiere['H' + str(k)]
                    grey_background(e)
                    grey_background(h)
                    k += 2

                k = 10
                while k < 29:
                    t = ws_matiere['A' + str(k)]
                    blue_background(t)
                    k += 2

                a28.border = Border(left=double, bottom=double)
                e28.border = Border(bottom=double)
                h28.border = Border(right=double, bottom=double)

                # FUSION DES CELLULES

                ws_matiere.merge_cells(start_row=1, start_column=1, end_row=4, end_column=4)
                ws_matiere.merge_cells(start_row=1, start_column=5, end_row=4, end_column=7)
                ws_matiere.merge_cells(start_row=1, start_column=8, end_row=2, end_column=9)
                ws_matiere.merge_cells(start_row=3, start_column=8, end_row=4, end_column=9)
                ws_matiere.merge_cells(start_row=5, start_column=8, end_row=6, end_column=9)
                ws_matiere.merge_cells(start_row=5, start_column=1, end_row=6, end_column=7)
                ws_matiere.merge_cells(start_row=7, start_column=1, end_row=8, end_column=4)
                ws_matiere.merge_cells(start_row=7, start_column=5, end_row=8, end_column=7)
                ws_matiere.merge_cells(start_row=7, start_column=8, end_row=8, end_column=9)

                def fusion(n):
                    ws_matiere.merge_cells(start_row=n, start_column=1, end_row=n, end_column=4)
                    ws_matiere.merge_cells(start_row=n, start_column=5, end_row=n, end_column=7)
                    ws_matiere.merge_cells(start_row=n, start_column=8, end_row=n, end_column=9)

                for i in range(9, 29):
                    fusion(i)

                # MISE EN PAGE DU TEXTE
                a1.font = Font(color="FFFFFF", bold=True, size=14)

                e1.font = Font(color="FFFFFF", bold=True, size=14)

                a7.font = Font(color="000000", bold=True, size=14)

                e7.font = Font(color="000000", bold=True, size=14)

                h7.font = Font(color="000000", bold=True, size=14)

                h1.font = Font(color="FFFFFF", bold=True, size=11)

                def titrer(cellule):
                    cellule.font = Font(color="FFFFFF", size=11)

                titrer(h3)
                titrer(h5)

                for i in range(9, 29):
                    t = ws_matiere['A' + str(i)]
                    titrer(t)
                    centrage(t)

                # fin calcule sur matiere____________________________________________________________________

                # calcule sur procedes_______________________________________________________________________
                wb.create_sheet('Procédes')
                ws_procedes = wb['Procédes']
                ws_procedes.title = "Procédés"
                                    
                #Remplir les cellule
                ws_procedes['A1'] = tab_procedes[0].a
                ws_procedes['A5'] = tab_procedes[2].a
                ws_procedes['A7'] = tab_procedes[3].a


                new = str(multiplicator)
                new_val = tab_procedes[0].e
                new_val = new_val.replace("1", new)
                ws_procedes['E1'] = new_val
                ws_procedes['E7'] = tab_procedes[3].e
        

                ws_procedes['H1'] = tab_procedes[0].h
                ws_procedes['H3'] = tab_procedes[1].h
                ws_procedes['H5'] = tab_procedes[2].h
                ws_procedes['H7'] = tab_procedes[3].h
                

                for i in range(9, 29):
                    x = i - 5           
                    ws_procedes['A'+str(i)]=tab_procedes[x].a
                    ws_procedes['H'+str(i)]=tab_procedes[x].h
                    do =(tab_procedes[i-5].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(tab_procedes[i-5].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(calc)
                    ws_procedes['E'+str(i)]=resultat
                    
                    
                    

                # DECLARATIONS DE QUELQUES CELLULES FREQUEMMENT UTILISEES
                a1 = ws_procedes['A1']
                e1 = ws_procedes['E1']
                a5 = ws_procedes['A5']
                a7 = ws_procedes['A7']
                e7 = ws_procedes['E7']
                h7 = ws_procedes['H7']
                h1 = ws_procedes['H1']
                h3 = ws_procedes['H3']
                h5 = ws_procedes['H5']
                a28 = ws_procedes['A28']
                e28 = ws_procedes['E28']
                h28 = ws_procedes['H28']

                # BORDURES ET ALIGNEMENT DES CELLULES
                thin = Side(border_style="thin", color="000000")
                double = Side(border_style="double", color="000000")

                def centrage(cellule):
                    cellule.alignment = Alignment(horizontal="center", vertical="center")

                def vertivaly_centrage(cellule):
                    cellule.alignment = Alignment(vertical="center")

                def blue_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="228B22")

                def bleuet_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="90EE90")

                def grey_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="DDDDDD")

                for i in range(9, 29):
                    j = ws_procedes['E' + str(i)]
                    centrage(j)

                for i in range(9, 29):
                    j = ws_procedes['H' + str(i)]
                    centrage(j)

                a1.border = Border(top=double, left=double)
                blue_background(a1)
                vertivaly_centrage(a1)

                e1.border = Border(top=double)
                blue_background(e1)
                centrage(e1)

                h1.border = Border(top=double, left=thin, right=double, bottom=thin)
                blue_background(h1)
                vertivaly_centrage(h1)

                h3.border = Border(top=thin, left=thin, right=double, bottom=thin)
                blue_background(h3)
                vertivaly_centrage(h3)

                h5.border = Border(top=thin, left=thin, right=double, bottom=double)
                blue_background(h5)
                vertivaly_centrage(h5)

                a7.border = Border(top=double, left=double, bottom=double)
                blue_background(a7)
                centrage(a7)

                e7.border = Border(top=double, bottom=double)
                blue_background(e7)
                centrage(e7)

                h7.border = Border(top=double, right=double, bottom=double)
                blue_background(h7)
                centrage(h7)

                a5.border = Border(left=double, bottom=double)
                blue_background(a5)
                vertivaly_centrage(a5)

                for i in range(9, 28):
                    t = ws_procedes['A' + str(i)]
                    t.border = Border(left=double)
                    t = ws_procedes['H' + str(i)]
                    t.border = Border(right=double)

                k = 9
                while k < 29:
                    t = ws_procedes['A' + str(k)]
                    bleuet_background(t)
                    e = ws_procedes['E' + str(k)]
                    h = ws_procedes['H' + str(k)]
                    grey_background(e)
                    grey_background(h)
                    k += 2

                k = 10
                while k < 29:
                    t = ws_procedes['A' + str(k)]
                    blue_background(t)
                    k += 2

                a28.border = Border(left=double, bottom=double)
                e28.border = Border(bottom=double)
                h28.border = Border(right=double, bottom=double)

                # FUSION DES CELLULES

                ws_procedes.merge_cells(start_row=1, start_column=1, end_row=4, end_column=4)
                ws_procedes.merge_cells(start_row=1, start_column=5, end_row=4, end_column=7)
                ws_procedes.merge_cells(start_row=1, start_column=8, end_row=2, end_column=9)
                ws_procedes.merge_cells(start_row=3, start_column=8, end_row=4, end_column=9)
                ws_procedes.merge_cells(start_row=5, start_column=8, end_row=6, end_column=9)
                ws_procedes.merge_cells(start_row=5, start_column=1, end_row=6, end_column=7)
                ws_procedes.merge_cells(start_row=7, start_column=1, end_row=8, end_column=4)
                ws_procedes.merge_cells(start_row=7, start_column=5, end_row=8, end_column=7)
                ws_procedes.merge_cells(start_row=7, start_column=8, end_row=8, end_column=9)

                def fusion(n):
                    ws_procedes.merge_cells(start_row=n, start_column=1, end_row=n, end_column=4)
                    ws_procedes.merge_cells(start_row=n, start_column=5, end_row=n, end_column=7)
                    ws_procedes.merge_cells(start_row=n, start_column=8, end_row=n, end_column=9)

                for i in range(9, 29):
                    fusion(i)

                # MISE EN PAGE DU TEXTE
                a1.font = Font(color="FFFFFF", bold=True, size=14)

                e1.font = Font(color="FFFFFF", bold=True, size=14)

                a7.font = Font(color="000000", bold=True, size=14)

                e7.font = Font(color="000000", bold=True, size=14)

                h7.font = Font(color="000000", bold=True, size=14)

                h1.font = Font(color="FFFFFF", bold=True, size=11)

                def titrer(cellule):
                    cellule.font = Font(color="FFFFFF", size=11)

                titrer(h3)
                titrer(h5)

                for i in range(9, 29):
                    t = ws_procedes['A' + str(i)]
                    titrer(t)
                    centrage(t)


                # fin calcule procedes :______________________________________________________________


                # calcule sur transport :_______________________________________________________________

                wb.create_sheet('Transport')
                ws_transport = wb['Transport']
                ws_transport.title = "Transport"
                                    
                #Remplir les cellule
                ws_transport['A1'] = tab_transport[0].a
                ws_transport['A5'] = tab_transport[2].a
                ws_transport['A7'] = tab_transport[3].a


                new = str(multip)
                new_val = tab_transport[0].e
                new_val = new_val.replace("1", new)
                ws_transport['E1'] = new_val
                ws_transport['E7'] = tab_transport[3].e
        

                ws_transport['H1'] = tab_transport[0].h
                ws_transport['H3'] = tab_transport[1].h
                ws_transport['H5'] = tab_transport[2].h
                ws_transport['H7'] = tab_transport[3].h
                

                for i in range(9, 29):
                    x = i - 5           
                    ws_transport['A'+str(i)]=tab_transport[x].a
                    ws_transport['H'+str(i)]=tab_transport[x].h
                    do =(tab_transport[i-5].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(tab_transport[i-5].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(ca)
                    ws_transport['E'+str(i)]=resultat
                    
                    
                    

                # DECLARATIONS DE QUELQUES CELLULES FREQUEMMENT UTILISEES
                a1 = ws_transport['A1']
                e1 = ws_transport['E1']
                a5 = ws_transport['A5']
                a7 = ws_transport['A7']
                e7 = ws_transport['E7']
                h7 = ws_transport['H7']
                h1 = ws_transport['H1']
                h3 = ws_transport['H3']
                h5 = ws_transport['H5']
                a28 = ws_transport['A28']
                e28 = ws_transport['E28']
                h28 = ws_transport['H28']

                # BORDURES ET ALIGNEMENT DES CELLULES
                thin = Side(border_style="thin", color="000000")
                double = Side(border_style="double", color="000000")

                def centrage(cellule):
                    cellule.alignment = Alignment(horizontal="center", vertical="center")

                def vertivaly_centrage(cellule):
                    cellule.alignment = Alignment(vertical="center")

                def blue_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="228B22")

                def bleuet_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="90EE90")

                def grey_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="DDDDDD")

                for i in range(9, 29):
                    j = ws_transport['E' + str(i)]
                    centrage(j)

                for i in range(9, 29):
                    j = ws_transport['H' + str(i)]
                    centrage(j)

                a1.border = Border(top=double, left=double)
                blue_background(a1)
                vertivaly_centrage(a1)

                e1.border = Border(top=double)
                blue_background(e1)
                centrage(e1)

                h1.border = Border(top=double, left=thin, right=double, bottom=thin)
                blue_background(h1)
                vertivaly_centrage(h1)

                h3.border = Border(top=thin, left=thin, right=double, bottom=thin)
                blue_background(h3)
                vertivaly_centrage(h3)

                h5.border = Border(top=thin, left=thin, right=double, bottom=double)
                blue_background(h5)
                vertivaly_centrage(h5)

                a7.border = Border(top=double, left=double, bottom=double)
                blue_background(a7)
                centrage(a7)

                e7.border = Border(top=double, bottom=double)
                blue_background(e7)
                centrage(e7)

                h7.border = Border(top=double, right=double, bottom=double)
                blue_background(h7)
                centrage(h7)

                a5.border = Border(left=double, bottom=double)
                blue_background(a5)
                vertivaly_centrage(a5)

                for i in range(9, 28):
                    t = ws_transport['A' + str(i)]
                    t.border = Border(left=double)
                    t = ws_transport['H' + str(i)]
                    t.border = Border(right=double)

                k = 9
                while k < 29:
                    t = ws_transport['A' + str(k)]
                    bleuet_background(t)
                    e = ws_transport['E' + str(k)]
                    h = ws_transport['H' + str(k)]
                    grey_background(e)
                    grey_background(h)
                    k += 2

                k = 10
                while k < 29:
                    t = ws_transport['A' + str(k)]
                    blue_background(t)
                    k += 2

                a28.border = Border(left=double, bottom=double)
                e28.border = Border(bottom=double)
                h28.border = Border(right=double, bottom=double)

                # FUSION DES CELLULES

                ws_transport.merge_cells(start_row=1, start_column=1, end_row=4, end_column=4)
                ws_transport.merge_cells(start_row=1, start_column=5, end_row=4, end_column=7)
                ws_transport.merge_cells(start_row=1, start_column=8, end_row=2, end_column=9)
                ws_transport.merge_cells(start_row=3, start_column=8, end_row=4, end_column=9)
                ws_transport.merge_cells(start_row=5, start_column=8, end_row=6, end_column=9)
                ws_transport.merge_cells(start_row=5, start_column=1, end_row=6, end_column=7)
                ws_transport.merge_cells(start_row=7, start_column=1, end_row=8, end_column=4)
                ws_transport.merge_cells(start_row=7, start_column=5, end_row=8, end_column=7)
                ws_transport.merge_cells(start_row=7, start_column=8, end_row=8, end_column=9)

                def fusion(n):
                    ws_transport.merge_cells(start_row=n, start_column=1, end_row=n, end_column=4)
                    ws_transport.merge_cells(start_row=n, start_column=5, end_row=n, end_column=7)
                    ws_transport.merge_cells(start_row=n, start_column=8, end_row=n, end_column=9)

                for i in range(9, 29):
                    fusion(i)

                # MISE EN PAGE DU TEXTE
                a1.font = Font(color="FFFFFF", bold=True, size=14)

                e1.font = Font(color="FFFFFF", bold=True, size=14)

                a7.font = Font(color="000000", bold=True, size=14)

                e7.font = Font(color="000000", bold=True, size=14)

                h7.font = Font(color="000000", bold=True, size=14)

                h1.font = Font(color="FFFFFF", bold=True, size=11)

                def titrer(cellule):
                    cellule.font = Font(color="FFFFFF", size=11)

                titrer(h3)
                titrer(h5)

                for i in range(9, 29):
                    t = ws_transport['A' + str(i)]
                    titrer(t)
                    centrage(t)

                # fin transport _-------------------------------------------------------------------------
                
                # calcule sur fin de vie :_______________________________________________________________

                wb.create_sheet('Scénario fin de vie')
                ws_fin_vie = wb['Scénario fin de vie']
                ws_fin_vie.title = "Scénario fin de vie"
                                    
                #Remplir les cellule
                ws_fin_vie['A1'] = tab_fin_vie[0].a
                ws_fin_vie['A5'] = tab_fin_vie[2].a
                ws_fin_vie['A7'] = tab_fin_vie[3].a


                new = str(multiplicator)
                new_val = tab_fin_vie[0].e
                new_val = new_val.replace("1", new)
                ws_fin_vie['E1'] = new_val
                ws_fin_vie['E7'] = tab_fin_vie[3].e
        

                ws_fin_vie['H1'] = tab_fin_vie[0].h
                ws_fin_vie['H3'] = tab_fin_vie[1].h
                ws_fin_vie['H5'] = tab_fin_vie[2].h
                ws_fin_vie['H7'] = tab_fin_vie[3].h
                

                for i in range(9, 29):
                    x = i - 5           
                    ws_fin_vie['A'+str(i)]=tab_fin_vie[x].a
                    ws_fin_vie['H'+str(i)]=tab_fin_vie[x].h
                    do =(tab_fin_vie[i-5].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(tab_fin_vie[i-5].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(calc)
                    ws_fin_vie['E'+str(i)]=resultat
                    
                    
                    

                # DECLARATIONS DE QUELQUES CELLULES FREQUEMMENT UTILISEES
                a1 = ws_fin_vie['A1']
                e1 = ws_fin_vie['E1']
                a5 = ws_fin_vie['A5']
                a7 = ws_fin_vie['A7']
                e7 = ws_fin_vie['E7']
                h7 = ws_fin_vie['H7']
                h1 = ws_fin_vie['H1']
                h3 = ws_fin_vie['H3']
                h5 = ws_fin_vie['H5']
                a28 = ws_fin_vie['A28']
                e28 = ws_fin_vie['E28']
                h28 = ws_fin_vie['H28']

                # BORDURES ET ALIGNEMENT DES CELLULES
                thin = Side(border_style="thin", color="000000")
                double = Side(border_style="double", color="000000")

                def centrage(cellule):
                    cellule.alignment = Alignment(horizontal="center", vertical="center")

                def vertivaly_centrage(cellule):
                    cellule.alignment = Alignment(vertical="center")

                def blue_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="228B22")

                def bleuet_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="90EE90")

                def grey_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="DDDDDD")

                for i in range(9, 29):
                    j = ws_fin_vie['E' + str(i)]
                    centrage(j)

                for i in range(9, 29):
                    j = ws_fin_vie['H' + str(i)]
                    centrage(j)

                a1.border = Border(top=double, left=double)
                blue_background(a1)
                vertivaly_centrage(a1)

                e1.border = Border(top=double)
                blue_background(e1)
                centrage(e1)

                h1.border = Border(top=double, left=thin, right=double, bottom=thin)
                blue_background(h1)
                vertivaly_centrage(h1)

                h3.border = Border(top=thin, left=thin, right=double, bottom=thin)
                blue_background(h3)
                vertivaly_centrage(h3)

                h5.border = Border(top=thin, left=thin, right=double, bottom=double)
                blue_background(h5)
                vertivaly_centrage(h5)

                a7.border = Border(top=double, left=double, bottom=double)
                blue_background(a7)
                centrage(a7)

                e7.border = Border(top=double, bottom=double)
                blue_background(e7)
                centrage(e7)

                h7.border = Border(top=double, right=double, bottom=double)
                blue_background(h7)
                centrage(h7)

                a5.border = Border(left=double, bottom=double)
                blue_background(a5)
                vertivaly_centrage(a5)

                for i in range(9, 28):
                    t = ws_fin_vie['A' + str(i)]
                    t.border = Border(left=double)
                    t = ws_fin_vie['H' + str(i)]
                    t.border = Border(right=double)

                k = 9
                while k < 29:
                    t = ws_fin_vie['A' + str(k)]
                    bleuet_background(t)
                    e = ws_fin_vie['E' + str(k)]
                    h = ws_fin_vie['H' + str(k)]
                    grey_background(e)
                    grey_background(h)
                    k += 2

                k = 10
                while k < 29:
                    t = ws_fin_vie['A' + str(k)]
                    blue_background(t)
                    k += 2

                a28.border = Border(left=double, bottom=double)
                e28.border = Border(bottom=double)
                h28.border = Border(right=double, bottom=double)

                # FUSION DES CELLULES

                ws_fin_vie.merge_cells(start_row=1, start_column=1, end_row=4, end_column=4)
                ws_fin_vie.merge_cells(start_row=1, start_column=5, end_row=4, end_column=7)
                ws_fin_vie.merge_cells(start_row=1, start_column=8, end_row=2, end_column=9)
                ws_fin_vie.merge_cells(start_row=3, start_column=8, end_row=4, end_column=9)
                ws_fin_vie.merge_cells(start_row=5, start_column=8, end_row=6, end_column=9)
                ws_fin_vie.merge_cells(start_row=5, start_column=1, end_row=6, end_column=7)
                ws_fin_vie.merge_cells(start_row=7, start_column=1, end_row=8, end_column=4)
                ws_fin_vie.merge_cells(start_row=7, start_column=5, end_row=8, end_column=7)
                ws_fin_vie.merge_cells(start_row=7, start_column=8, end_row=8, end_column=9)

                def fusion(n):
                    ws_fin_vie.merge_cells(start_row=n, start_column=1, end_row=n, end_column=4)
                    ws_fin_vie.merge_cells(start_row=n, start_column=5, end_row=n, end_column=7)
                    ws_fin_vie.merge_cells(start_row=n, start_column=8, end_row=n, end_column=9)

                for i in range(9, 29):
                    fusion(i)

                # MISE EN PAGE DU TEXTE
                a1.font = Font(color="FFFFFF", bold=True, size=14)

                e1.font = Font(color="FFFFFF", bold=True, size=14)

                a7.font = Font(color="000000", bold=True, size=14)

                e7.font = Font(color="000000", bold=True, size=14)

                h7.font = Font(color="000000", bold=True, size=14)

                h1.font = Font(color="FFFFFF", bold=True, size=11)

                def titrer(cellule):
                    cellule.font = Font(color="FFFFFF", size=11)

                titrer(h3)
                titrer(h5)

                for i in range(9, 29):
                    t = ws_fin_vie['A' + str(i)]
                    titrer(t)
                    centrage(t)

                # fin de calcule fin_de_vie____________________________________________________________________
               
                
                # calcule sur energie :_______________________________________________________________

                wb.create_sheet('Energie')
                ws_energie = wb['Energie']
                ws_energie .title = "Energie"
                                    
                #Remplir les cellule
                ws_energie['A1'] = tab_energie[0].a
                ws_energie['A5'] = tab_energie[2].a
                ws_energie['A7'] = tab_energie[3].a


                new = str(multiplicator)
                new_val = tab_energie[0].e
                new_val = new_val.replace("1", new)
                ws_energie['E1'] = new_val
                ws_energie['E7'] = tab_energie[3].e
        

                ws_energie['H1'] = tab_energie[0].h
                ws_energie['H3'] = tab_energie[1].h
                ws_energie['H5'] = tab_energie[2].h
                ws_energie['H7'] = tab_energie[3].h
                

                for i in range(9, 29):
                    x = i - 5           
                    ws_energie['A'+str(i)]=tab_energie[x].a
                    ws_energie['H'+str(i)]=tab_energie[x].h
                    do =(tab_energie[i-5].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(tab_energie[i-5].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(calc)
                    ws_energie['E'+str(i)]=resultat
                    
                    
                    

                # DECLARATIONS DE QUELQUES CELLULES FREQUEMMENT UTILISEES
                a1 = ws_energie['A1']
                e1 = ws_energie['E1']
                a5 = ws_energie['A5']
                a7 = ws_energie['A7']
                e7 = ws_energie['E7']
                h7 = ws_energie['H7']
                h1 = ws_energie['H1']
                h3 = ws_energie['H3']
                h5 = ws_energie['H5']
                a28 = ws_energie['A28']
                e28 = ws_energie['E28']
                h28 = ws_energie['H28']

                # BORDURES ET ALIGNEMENT DES CELLULES
                thin = Side(border_style="thin", color="000000")
                double = Side(border_style="double", color="000000")

                def centrage(cellule):
                    cellule.alignment = Alignment(horizontal="center", vertical="center")

                def vertivaly_centrage(cellule):
                    cellule.alignment = Alignment(vertical="center")

                def blue_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="228B22")

                def bleuet_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="90EE90")

                def grey_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="DDDDDD")

                for i in range(9, 29):
                    j = ws_energie['E' + str(i)]
                    centrage(j)

                for i in range(9, 29):
                    j = ws_energie['H' + str(i)]
                    centrage(j)

                a1.border = Border(top=double, left=double)
                blue_background(a1)
                vertivaly_centrage(a1)

                e1.border = Border(top=double)
                blue_background(e1)
                centrage(e1)

                h1.border = Border(top=double, left=thin, right=double, bottom=thin)
                blue_background(h1)
                vertivaly_centrage(h1)

                h3.border = Border(top=thin, left=thin, right=double, bottom=thin)
                blue_background(h3)
                vertivaly_centrage(h3)

                h5.border = Border(top=thin, left=thin, right=double, bottom=double)
                blue_background(h5)
                vertivaly_centrage(h5)

                a7.border = Border(top=double, left=double, bottom=double)
                blue_background(a7)
                centrage(a7)

                e7.border = Border(top=double, bottom=double)
                blue_background(e7)
                centrage(e7)

                h7.border = Border(top=double, right=double, bottom=double)
                blue_background(h7)
                centrage(h7)

                a5.border = Border(left=double, bottom=double)
                blue_background(a5)
                vertivaly_centrage(a5)

                for i in range(9, 28):
                    t = ws_energie['A' + str(i)]
                    t.border = Border(left=double)
                    t = ws_energie['H' + str(i)]
                    t.border = Border(right=double)

                k = 9
                while k < 29:
                    t = ws_energie['A' + str(k)]
                    bleuet_background(t)
                    e = ws_energie['E' + str(k)]
                    h = ws_energie['H' + str(k)]
                    grey_background(e)
                    grey_background(h)
                    k += 2

                k = 10
                while k < 29:
                    t = ws_energie['A' + str(k)]
                    blue_background(t)
                    k += 2

                a28.border = Border(left=double, bottom=double)
                e28.border = Border(bottom=double)
                h28.border = Border(right=double, bottom=double)

                # FUSION DES CELLULES

                ws_energie.merge_cells(start_row=1, start_column=1, end_row=4, end_column=4)
                ws_energie.merge_cells(start_row=1, start_column=5, end_row=4, end_column=7)
                ws_energie.merge_cells(start_row=1, start_column=8, end_row=2, end_column=9)
                ws_energie.merge_cells(start_row=3, start_column=8, end_row=4, end_column=9)
                ws_energie.merge_cells(start_row=5, start_column=8, end_row=6, end_column=9)
                ws_energie.merge_cells(start_row=5, start_column=1, end_row=6, end_column=7)
                ws_energie.merge_cells(start_row=7, start_column=1, end_row=8, end_column=4)
                ws_energie.merge_cells(start_row=7, start_column=5, end_row=8, end_column=7)
                ws_energie.merge_cells(start_row=7, start_column=8, end_row=8, end_column=9)

                def fusion(n):
                    ws_energie.merge_cells(start_row=n, start_column=1, end_row=n, end_column=4)
                    ws_energie.merge_cells(start_row=n, start_column=5, end_row=n, end_column=7)
                    ws_energie.merge_cells(start_row=n, start_column=8, end_row=n, end_column=9)

                for i in range(9, 29):
                    fusion(i)

                # MISE EN PAGE DU TEXTE
                a1.font = Font(color="FFFFFF", bold=True, size=14)

                e1.font = Font(color="FFFFFF", bold=True, size=14)

                a7.font = Font(color="000000", bold=True, size=14)

                e7.font = Font(color="000000", bold=True, size=14)

                h7.font = Font(color="000000", bold=True, size=14)

                h1.font = Font(color="FFFFFF", bold=True, size=11)

                def titrer(cellule):
                    cellule.font = Font(color="FFFFFF", size=11)

                titrer(h3)
                titrer(h5)

                for i in range(9, 29):
                    t = ws_energie['A' + str(i)]
                    titrer(t)
                    centrage(t)

                
                
                
                # racp_________________________________________________________________________________

                wb.create_sheet('Récapitulatif')
                ws_recap = wb['Récapitulatif']
                ws_recap.title = "Récapitulatif"
                                    
                #Remplir les cellule
                ws_recap['A1'] = "Critères"
                ws_recap['A2'] = "Matière"
                ws_recap['A3'] = "Procédé de fabrication"
                ws_recap['A4'] = "Moyen de transport"
                ws_recap['A5'] = "Scénario de fin de vie"
                ws_recap['A6'] = "Poids(En Kg)"

                ws_recap['E1'] = "Valeurs"
                ws_recap['E2'] = tab_matiere[0].a
                ws_recap['E3'] = tab_procedes[0].a
                ws_recap['E4'] = tab_transport[0].a
                ws_recap['E5'] = tab_fin_vie[0].a
                ws_recap['E6'] = multiplicator

                
                ws_recap['E7'] = ws_recap['E13'] = ws_recap['E19']= "MJ"
                ws_recap['A8'] = ws_recap['A14'] = ws_recap['A20']= "Matière"
                ws_recap['A9'] = ws_recap['A15'] = ws_recap['A21']= "Procédé"
                ws_recap['A10'] = ws_recap['A16'] = ws_recap['A22']= "Transport"
                ws_recap['A11'] = ws_recap['A17'] = ws_recap['A23']= "Fin de vie" 
                ws_recap['A12'] = ws_recap['A18'] = ws_recap['A24']= "Total"  


                ws_recap['A7'] = "Energie primaire totale"
                ws_recap['A13'] = "Energie renouvelable"
                ws_recap['A19'] = "Energie non renouvelable"

                      
                def calcresultmatiere(n):      
                    do =(tab_matiere[n].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(tab_matiere[n].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(calc)
                    return resultat

                

                def vertivaly_left(cellule):
                    cellule.alignment = Alignment(horizontal="left", vertical="center")
                      
                ws_recap['E8']=calcresultmatiere(16)
                ws_recap['E14']=calcresultmatiere(17)
                ws_recap['E20']=calcresultmatiere(18)

                

                for i in range(8,25):
                    vertivaly_left(ws_recap['E'+str(i)])
                # vertivaly_left(ws_recap['E14'])
                # vertivaly_left(ws_recap['E20'])

                
                def calcresultprocedes(n):      
                    do =(tab_procedes[n].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(tab_procedes[n].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(calc)
                    return resultat

                ws_recap['E9']=calcresultprocedes(16)
                ws_recap['E15']=calcresultprocedes(17)
                ws_recap['E21']=calcresultprocedes(18)

                def calcresulttransport(n):      
                    do =(tab_transport[n].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(tab_transport[n].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(ca)
                    return resultat

                ws_recap['E10']=calcresulttransport(16)
                ws_recap['E16']=calcresulttransport(17)
                ws_recap['E22']=calcresulttransport(18)
                

                def calcresultfin(n):      
                    do =(tab_fin_vie[n].e[0:4])
                    doo = do.replace(",",".")  
                    puissance =int(tab_fin_vie[n].e[5:8])
                    don = float(doo)* pow(10,puissance)
                    resultat = float(don) * float(calc)
                    return resultat

                ws_recap['E11']=calcresultfin(16)
                ws_recap['E17']=calcresultfin(17)
                ws_recap['E23']=calcresultfin(18)

                
                totale1 = float(calcresultmatiere(16))+float(calcresultprocedes(16))+float(calcresulttransport(16)+float(calcresultfin(16)))
                totale2 = float(calcresultmatiere(17))+float(calcresultprocedes(17))+float(calcresulttransport(17)+float(calcresultfin(17)))
                totale3 = float(calcresultmatiere(18))+float(calcresultprocedes(18))+float(calcresulttransport(18)+float(calcresultfin(18)))

                ws_recap['E12'] = totale1
                ws_recap['E18'] = totale2
                ws_recap['E24'] = totale3
                

                def green_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="228B22")

                def greeeen_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="90EE90")

                def grey_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="DDDDDD")

                def neww_background(cellule):
                    cellule.fill = PatternFill("solid", fgColor="729179")

                green_background(ws_recap['A1'])
                green_background(ws_recap['E1'])
                green_background(ws_recap['A7'])
                green_background(ws_recap['E7'])
                green_background(ws_recap['A13'])
                green_background(ws_recap['E13'])
                green_background(ws_recap['A19'])
                green_background(ws_recap['E19'])
                

                ws_recap['A12'].font = Font(color="FFFFFF", bold=True, size=14)
                ws_recap['A18'].font = Font(color="FFFFFF", bold=True, size=14)
                ws_recap['A24'].font = Font(color="FFFFFF", bold=True, size=14)

                ws_recap['A7'].font = Font( bold=True, size=14)
                ws_recap['A13'].font = Font( bold=True, size=14)
                ws_recap['A19'].font = Font(bold=True, size=14)

                ws_recap['E7'].font = Font(bold=True, size=14)
                ws_recap['E13'].font = Font(bold=True, size=14)
                ws_recap['E19'].font = Font( bold=True, size=14)



                neww_background(ws_recap['A12'])
                neww_background(ws_recap['E12'])
                neww_background(ws_recap['A18'])
                neww_background(ws_recap['E18'])
                neww_background(ws_recap['A24'])
                neww_background(ws_recap['E24'])


                for i in range(2,7):
                    greeeen_background(ws_recap['A'+str(i)])
                    greeeen_background(ws_recap['E'+str(i)])

                for i in range(8,12):
                    grey_background(ws_recap['A'+str(i)])
                
                for i in range(14,18):
                    grey_background(ws_recap['A'+str(i)])
                
                for i in range(20,24):
                    grey_background(ws_recap['A'+str(i)])
                
                # grey_background(ws_recap['E12'])
                # grey_background(ws_recap['E18'])
                # grey_background(ws_recap['E24'])

                ws_recap['A1'].font = Font(color="FFFFFF", bold=True, size=14)
                ws_recap['E1'].font = Font(color="FFFFFF", bold=True, size=14)

                

                def fusion(n):
                    ws_recap.merge_cells(start_row=n, start_column=1, end_row=n, end_column=4)
                    ws_recap.merge_cells(start_row=n, start_column=5, end_row=n, end_column=8)

                for i in range(1, 25):
                    fusion(i)
                
          

                # fin recap_________________________________________________________________________________
                

                # bilan note unique_________________________________________________________________________________

                wb.create_sheet('Bilan Note Unique')
                ws_note = wb['Bilan Note Unique']
                ws_note.title = "Bilan Note Unique"
                                    
                #Remplir les cellule
                ws_note['A1'] = "Note unique"
                ws_note['A2'] = "Matière"
                ws_note['A3'] = "Procédé de fabrication"
                ws_note['A4'] = "Moyen de transport"
                ws_note['A5'] = "Scénario de fin de vie"
                ws_note['A6'] = "Energie"
                ws_note['A7'] = "Total"
                ws_note['E1'] = "Points"

                ws_note['E2']=calcresultmatiere(23)
                ws_note['E3']=calcresultprocedes(23)
                ws_note['E4']=calcresulttransport(23)
                ws_note['E5']=calcresultfin(23)

                do =(tab_energie[23].e[0:4])
                doo = do.replace(",",".")  
                puissance =int(tab_energie[23].e[5:8])
                don = float(doo)* pow(10,puissance)
                note_energie = float(don) * float(calc)
                ws_note['E6']=note_energie

                totale=float(note_energie)+float(calcresultmatiere(23))+float(calcresultprocedes(23))+float(calcresulttransport(23))+float(calcresultfin(23))
                ws_note['E7']=totale
                    

                for i in range(2,8):
                    vertivaly_left(ws_note['E'+str(i)])

                for i in range(2,7):
                    grey_background(ws_note['A'+str(i)])

                green_background(ws_note['A1'])
                green_background(ws_note['E1'])
                neww_background(ws_note['A7'])
                neww_background(ws_note['E7'])

                def fusion(n):
                    ws_note.merge_cells(start_row=n, start_column=1, end_row=n, end_column=4)
                    ws_note.merge_cells(start_row=n, start_column=5, end_row=n, end_column=8)
                for i in range(1,8):
                    fusion(i)

                ws_note['A1'].font = Font(color="FFFFFF", bold=True, size=14)
                ws_note['E1'].font = Font(color="FFFFFF", bold=True, size=14)

                # fin bilan note unique_____________________________________________________ 

                wb.save('/Users/lenovo/Desktop/'+nom_fich+'.xlsx')

                messages.success(request,"Votre document excel a été généré avec succes")
                return render(request,'calcules/recap.html',{'susses':'alert-success'})  

        else:
            messages.error(request,"il faut saisir un nombre dans les champs du quantité de matière et distance")
            return render(request,'calcules/recap.html',{'danger':'alert-danger'}) 
    return render(request,'calcules/recap.html') 



